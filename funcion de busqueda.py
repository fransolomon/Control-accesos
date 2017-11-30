from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

idnumber= "1237"

def buscar_nombre_del_usuario(id_numero):
	nombre_usuario=[]
	dest_filename = 'DATOSENTRADASALIDA.xlsx'
	wb = load_workbook(dest_filename, read_only=True)
	ws2 = wb[wb.sheetnames[1]] # seleccionamos la hoja 2 del excel. 

	for i in range(0,int(ws2.max_row)): # con el size maximo de la hoja de excell empezamos un bucle for de busqueda. por la columna de ID.
		cell = ws2.cell(row=i+1, column=1).value
		if (cell == id_numero):
			nombre_usuario = ws2.cell(row=i+1, column=2).value 
			break
	return nombre_usuario

print "El nombre del usuario para la id: ", idnumber, "seria :", buscar_nombre_del_usuario(idnumber)
quit()





# dest_filename = 'DATOSENTRADASALIDA.xlsx'
# wb = load_workbook(dest_filename, read_only=True)
# ws2 = wb[wb.sheetnames[1]] # entramos en la 

# print "max: ",ws2.max_row
# print "max: ",ws2.max_column

# for i in range(0,int(ws2.max_row)):
# 	cell = ws2.cell(row=i+1, column=1).value
# 	print "F: ",i, " - ", cell
# 	if (cell == 'FRAN'):
# 		print "ENCONTRADO numero de ID: ", ws2.cell(row=i+1, column=2).value 
# 		break



#La siguiente funcion busca en la hoja de datos de excel el ID del usuario y Retorna su nombre.

