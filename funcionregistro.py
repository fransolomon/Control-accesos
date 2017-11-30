from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import time

idnumber= "1237"
dest_filename = 'DATOSENTRADASALIDA.xlsx'

def buscar_nombre_del_usuario(id_numero):
	nombre_usuario=[]
	dest_filename = 'DATOSENTRADASALIDA.xlsx'
	wb = load_workbook(dest_filename, read_only=True)
	ws2 = wb[wb.sheetnames[1]] # seleccionamos la hoja 2 del excel. 

	for i in range(0,int(ws2.max_row)): # con el size maximo de la hoja de excell empezamos un bucle for de busqueda. por la columna de ID.
		cell = ws2.cell(row=i+1, column=1).value
		if (cell == id_numero):
			nombre_usuario = ws2.cell(row=i+1, column=2).value 
			break
	return nombre_usuario


def entrada_al_registro(id_numero):
	nombre_usuario=[]
	nombre_usuario = buscar_nombre_del_usuario(idnumber)
	
	dest_filename = 'DATOSENTRADASALIDA.xlsx'
	wb = load_workbook(dest_filename)
	ws2 = wb[wb.sheetnames[0]] # seleccionamos la hoja 1 del excel entradas. 
	print "[user: ",nombre_usuario,"] - En esta hoja el numero maximo del cols: ",ws2.max_row
	max_fila = ws2.max_row
	Hora = time.strftime("%H:%M:%S")
	Fecha =  time.strftime("%d/%m/%y")
	Estado = "nada"
	Notas = "nada"
	ws2.cell(row=max_fila+1, column=1, value=Fecha)
	ws2.cell(row=max_fila+1, column=2, value=Hora)
	ws2.cell(row=max_fila+1, column=3, value=Estado)
	ws2.cell(row=max_fila+1, column=4, value=id_numero)
	ws2.cell(row=max_fila+1, column=5, value=nombre_usuario)
	ws2.cell(row=max_fila+1, column=6, value=Notas)
	wb.save(filename = dest_filename)
	return

print "El nombre del usuario para la id: ", idnumber, "seria :", buscar_nombre_del_usuario(idnumber)
entrada_al_registro(idnumber)
quit()





# dest_filename = 'DATOSENTRADASALIDA.xlsx'
# wb = load_workbook(dest_filename, read_only=True)
# ws2 = wb[wb.sheetnames[1]] # entramos en la 

# print "max: ",ws2.max_row
# print "max: ",ws2.max_column

# for i in range(0,int(ws2.max_row)):
# 	cell = ws2.cell(row=i+1, column=1).value
# 	print "F: ",i, " - ", cell
# 	if (cell == 'FRAN'):
# 		print "ENCONTRADO numero de ID: ", ws2.cell(row=i+1, column=2).value 
# 		break



#La siguiente funcion busca en la hoja de datos de excel el ID del usuario y Retorna su nombre.

