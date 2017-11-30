#from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
dest_filename = 'DATOSENTRADASALIDA.xlsx'
#wb = load_workbook(dest_filename)
wb = Workbook()
ws1 = wb.active
ws1.title = "Registro movimientos"
ws1["A1"] = "FECHA"
ws1["B1"] = "HORA"
ws1["C1"] = "ESTADO"
ws1["D1"] = "IDNUM"
ws1["E1"] = "NOMBRE USUARIO"
ws1["F1"] = "NOTAS"
#ws1.cell(row=3, column=4, value="TEST"+a+a)

ws2 = wb.create_sheet(title="NOMBRESVSID")
ws2['A1'] = "ID"
ws2['B1'] = "NOMBRE"
ws2['B2'] = "PACO"
ws2['B3'] = "TOTE"
ws2['B4'] = "ASDAS"
ws2['B5'] = "ASDAS"
ws2['B6'] = "FRANCISCO"

ws2['A2'] = "1238"
ws2['A3'] = "1237"
ws2['A4'] = "1236"
ws2['A5'] = "1235"
ws2['A6'] = "C64BFCC4B5"
try:
	wb.save(filename = dest_filename)
except:
	print "el archivo de excel no consigue grabarse."