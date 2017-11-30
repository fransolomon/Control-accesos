#!usr/bin/python
import getpass
import sys
import telnetlib
import time
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
#import xlwt
# ....
# funciona OK en linux sin modificaciones. Tambien tiene una serie de cuestiones a mejorar:
# 1) Un tiempo para ir comprobando la señal de control del aparato.
# 2) La lista de usuarios completa en el PDF o que el la vaya generando si le da el OK la puerta.
# ....

HOST = "10.10.0.159"
PORT=80
TIMEOUT=10


def buscar_nombre_del_usuario(id_numero):
	nombre_usuario=[]
	dest_filename = 'DATOSENTRADASALIDA.xlsx'
	wb = load_workbook(dest_filename, read_only=True)
	ws2 = wb[wb.sheetnames[1]] # seleccionamos la hoja 2 del excel. 

	for i in range(0,int(ws2.max_row)): # con el size maximo de la hoja de excell empezamos un bucle for de busqueda. por la columna de ID.
		cell = ws2.cell(row=i+1, column=1).value
		print "--------------"
		print str(cell)," = ",str(id_numero)
		print "--------------"
		if (str(cell) in str(id_numero)):
			nombre_usuario = ws2.cell(row=i+1, column=2).value 
			return nombre_usuario
	print "NO EN LISTA"
	return "NO EN LISTA"

def entrada_al_registro(id_numero,Notas,Estado):
	nombre_usuario=[]

	nombre_usuario = buscar_nombre_del_usuario(id_numero)
	print id_numero
	dest_filename = 'DATOSENTRADASALIDA.xlsx'
	wb = load_workbook(dest_filename)
	ws2 = wb[wb.sheetnames[0]] # seleccionamos la hoja 1 del excel entradas. 
	print "[user: ",nombre_usuario,"] - En esta hoja el numero maximo del cols: ",ws2.max_row
	max_fila = ws2.max_row
	Hora = time.strftime("%H:%M:%S")
	Fecha =  time.strftime("%d/%m/%y")
	id_numero =str(id_numero)
	Notas = str(Notas)
	ws2.cell(row=max_fila+1, column=1, value=Fecha)
	ws2.cell(row=max_fila+1, column=2, value=Hora)
	ws2.cell(row=max_fila+1, column=3, value=Estado)
	ws2.cell(row=max_fila+1, column=4, value=id_numero)
	ws2.cell(row=max_fila+1, column=5, value=nombre_usuario)
	ws2.cell(row=max_fila+1, column=6, value=Notas)
	wb.save(filename = dest_filename)
	return

def main():
	try:
		print "Conectando con direccion IP:", HOST, "Al puerto: ",PORT
		t = telnetlib.Telnet()
		t.close()
		t.open('10.10.0.159', port=80)
		print "Conectando con direccion IP:", HOST, "Al puerto: ",PORT
		while True:
			a=t.read_until("HOLA",1)
			#print "En espera ... "
			if ("Control" in a):
				#t.close()
				Tiempo_de_inicio = time.strftime("%H:%M:%S")
				Fecha =  time.strftime("%d/%m/%y")
				print "Se recibio un dato de control [",Fecha,",",Tiempo_de_inicio,"]: ", a
			elif("Abre" in a): #ex. C64BFCC4B5 Abre
				Tiempo_de_inicio = time.strftime("%H:%M:%S")
				Fecha =  time.strftime("%d/%m/%y")
				buff = "Apertura de puerta del usuario ",a[0:11]," [",Fecha,",",Tiempo_de_inicio,"]: ", a
				print buff
				entrada_al_registro(a[0:11],buff,"Abre")
			elif("Maestra" in a): #ex. 1E02420759 Abre 1E02420759 Maestra 
				Tiempo_de_inicio = time.strftime("%H:%M:%S")
				Fecha =  time.strftime("%d/%m/%y")
				buff = "LLave maestra ",a[0:11]," [",Fecha,",",Tiempo_de_inicio,"]: ", a
				print buff
				entrada_al_registro(a[0:11],buff,"Maestra")
			elif("Negra" in a): #ex. 3B4BFCD05C Negra 
				Tiempo_de_inicio = time.strftime("%H:%M:%S")
				Fecha =  time.strftime("%d/%m/%y")
				buff = "LLave en la lista negra del lector ",a[0:11]," [",Fecha,",",Tiempo_de_inicio,"]: ", a
				print buff
				entrada_al_registro(a[0:11],buff,"Negra")
			elif("Cerrado" in a): #ex. 6A4BFCE439 Cerrado
				Tiempo_de_inicio = time.strftime("%H:%M:%S")
				Fecha =  time.strftime("%d/%m/%y")
				buff = "Se intenta introducir una llave fuera de lista en el lector",a[0:11]," [",Fecha,",",Tiempo_de_inicio,"]: ", a
				print buff
				entrada_al_registro(a[0:11],buff,"Cerrado")
		print "FIN DEL PROGRAMA"

	except:
	 	print "fail connection..."
	 	pass

if __name__ == '__main__':
    main()
#1E02420759 Abre 1E02420759 Maestra 
#3B4BFCD05C Negra 
#6A4BFCE439 Cerrado
# import os
 
# hostname = "lawebdelprogramxador.com"
# # response = os.system("ping -c 1 " + hostname) -> este seria para Windows
# response = os.system("ping -c 1 " + hostname + " > /dev/null 2>&1")
 
 
# if response == 0:
#     print ("%s responde" % hostname)
# else:
#     print ("%s no responde" % hostname)